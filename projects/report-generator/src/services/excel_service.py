import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from src.models.report_request import ReportRequest


class ExcelReportService:
    def generate(self, request: ReportRequest) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Reporte'

        header_fill = PatternFill('solid', fgColor='6366F1')
        header_font = Font(color='FFFFFF', bold=True)

        sheet['A1'] = request.title
        sheet['A1'].font = Font(size=14, bold=True, color='6366F1')
        sheet['A2'] = f'Autor: {request.author}'

        headers = ['Categoría', 'Concepto', 'Valor']
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=4, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        for row_index, item in enumerate(request.items, start=5):
            sheet.cell(row=row_index, column=1, value=item.category)
            sheet.cell(row=row_index, column=2, value=item.label)
            sheet.cell(row=row_index, column=3, value=item.value)

        sheet.column_dimensions['A'].width = 18
        sheet.column_dimensions['B'].width = 35
        sheet.column_dimensions['C'].width = 14

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.read()
